import os
import re
from itertools import combinations

import streamlit
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

import src.parsing as parsing


#######################
#     COORDINATES     #
#######################


def get_copositivity_analysis_coordinates(analysis_type, n_channels):
    analysis_coordinates = {
        'Count': {
            2: (19, 23),
            3: (27, 34),
            4: (35, 49),
            5: (43, 72)},
        'Area': {
            2: (13, 17),
            3: (18, 25),
            4: (23, 37),
            5: (28, 57)}
    }

    return analysis_coordinates[analysis_type][n_channels]


def get_copositivity_coordinates(n_channels):

    coordinates = {
        2: (10, 12),
        3: (11, 16),
        4: (12, 25),
        5: (13, 40)
    }

    return coordinates.get(n_channels, None)


def get_summary_coordinates(n_channels):
    coordinates = {
        2: (10, 11),
        3: (14, 18),
        4: (18, 29),
        5: (22, 48)
    }

    return coordinates.get(n_channels, None)


########################
# CHANNEL COMBINATIONS #
########################

def combine_channels(channels) -> list:
    """Returns a list of all possible channels combinations"""

    channels_combinations = []
    for i in range(2, len(channels) + 1):
        comb = [list(i) for i in list(combinations(channels, i))]
        channels_combinations = channels_combinations + comb
    return channels_combinations


def get_combination_name(channels_combinations):
    return '+'.join(str(channel) for channel in channels_combinations)


def reformat_channel(channels_dict):
    return {channel_name: re.split(r'[()]', channel_number)[1] for channel_number, channel_name in channels_dict.items()}


########################
#  PARSING/FORMATTING  #
########################


def rename_copositivity_columns(ws, channels, start_col):

    reformatted_channels = reformat_channel(channels)
    combinations_columns = combine_channels(reformatted_channels.values())

    for col, chan_combinations in enumerate(combinations_columns, start=start_col):
        col_name = get_combination_name(chan_combinations)
        ws.cell(2, col).value = col_name


def rename_summary_copositivity_columns(ws, file_channels, start_row):

    channels_in_file = {}
    for sample, channels in file_channels.items():
        for channel, channel_name in channels.items():
            channels_in_file[channel] = channel_name

    reformatted_channels = reformat_channel(channels_dict=channels_in_file)
    combinations_columns = combine_channels(reformatted_channels.values())

    for col, chan_combinations in enumerate(combinations_columns, start=start_row):
        col_name = get_combination_name(chan_combinations)
        ws.cell(1, col).value = col_name


def rename_copositivity_rows(ws, channels, start_row):
    reformatted_channels = reformat_channel(channels)
    combinations_columns = combine_channels(reformatted_channels.values())

    for row, chan_combinations in enumerate(combinations_columns, start=start_row):
        col_name = get_combination_name(chan_combinations)
        ws.cell(row, 1).value = col_name


def add_conditional_formatting(ws, col_start, col_end, row_start, row_end):
    columns_range = f'{col_start}{row_start}:{col_end}{row_end}'
    format_rule = ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='ff5370')
    ws.conditional_formatting.add(columns_range, format_rule)


def parse_copositivity_analysis_template(template_ws, destination_ws, analysis_type, n_channels):
    start_row, end_row = get_copositivity_analysis_coordinates(analysis_type, n_channels)

    parsing.copy_from_template(template_ws=template_ws, destination_ws=destination_ws,
                               start_row=start_row, end_row=end_row,
                               template_start_col=1, template_end_col=6,
                               destination_start_col=1, destination_end_col=6,
                               copy_value=True, copy_style=True)

    return


#######################
#        MAIN         #
#######################

def parse_copositivity_template(writer, sheets, filename, file_channels, analysis_end, analysis_type,
                                template_file: str, progress_bar=None):
    """ Copy template to add co-positivity columns """

    # Initialize progress_bar
    step = 0
    steps = len(sheets)
    if progress_bar:
        progress_bar = streamlit.progress(step, f'Parsing co-positivity template: [{step}/{steps}]')

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), template_file)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    workbook = writer.book

    for sheet in sheets:

        destination_ws = writer.sheets[sheet]

        # get last row
        end_row = destination_ws.max_row

        # find number of channels used to determine columns to parse
        channels = file_channels[sheet]
        n_channels = len(channels)

        if n_channels < 2:
            step += 1
            progress_bar.progress(step / steps, f'Parsing co-positivity template: [{step}/{steps}]')
            continue

        col_start, col_end = get_copositivity_coordinates(n_channels)

        parsing.copy_from_template(template_ws=template_ws, destination_ws=destination_ws,
                                   start_row=1, end_row=end_row,
                                   template_start_col=col_start, template_end_col=col_end,
                                   destination_start_col=col_start, destination_end_col=col_end,
                                   copy_value=True, copy_style=True)

        # rename columns
        rename_copositivity_columns(ws=destination_ws, channels=channels, start_col=col_start + 1)

        # add conditional formating
        add_conditional_formatting(ws=destination_ws, row_start=4, row_end=end_row,
                                   col_start=get_column_letter(col_start + 1),
                                   col_end=get_column_letter(col_end - 1))

        # parse copositivity analysis
        parse_copositivity_analysis_template(template_ws=template_ws, destination_ws=destination_ws,
                                             analysis_type=analysis_type, n_channels=n_channels)

        # rename_rows
        rename_copositivity_rows(ws=destination_ws, channels=channels, start_row=analysis_end[sheet] + 3)

        if progress_bar:
            step += 1
            progress_bar.progress(step / steps, f'Parsing co-positivity template: [{step}/{steps}]')

    workbook.save(filename)
    return


def parse_copositivity_summary(writer, filename, n_channels, file_channels, summary_template: str):

    # open summary sheet
    workbook = writer.book
    summary_ws = writer.sheets['summary']

    max_rowdata = summary_ws.max_row
    start, end = get_summary_coordinates(n_channels)

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), summary_template)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    # copy template analysis
    parsing.copy_from_template(template_ws, summary_ws,
                               start_row=1, end_row=max_rowdata,
                               template_start_col=start, template_end_col=end,
                               destination_start_col=start, destination_end_col=end)

    rename_summary_copositivity_columns(ws=summary_ws, file_channels=file_channels, start_row=start)

    # save file
    workbook.save(filename)
    return
