import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension


#######################
#        COLUMNS      #
#######################

def resize_analysis_columns(writer, sheets, n_channels, copositivity=False):
    """ Resize columns width for each .xlsx sheets. """

    for sheet in sheets:

        destination_ws = writer.sheets[sheet]
        maxcol = 7 + n_channels

        dim_holder = DimensionHolder(worksheet=destination_ws)

        # Analysis columns : From columns A (col = 1) to E (col = 5)
        for col in range(1, 6):
            if col == 1:
                dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=20)
            else:
                dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=15)

        # Separator colum : always column F (col = 6)
        dim_holder[get_column_letter(6)] = ColumnDimension(destination_ws, min=6, max=6, width=5)

        # Image data columns : From column G (col = 7) to last channel column (col = max_coldata - 1)
        for col in range(7, maxcol):
            if col == 7:
                dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=12)
            else:
                dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=16)

        # Separator colum : variable column based on number of channels (col = max_coldata)
        dim_holder[get_column_letter(maxcol)] = ColumnDimension(destination_ws, min=maxcol, max=maxcol, width=5)

        # Modify copositivity columns
        if copositivity:

            ...

        # Store modifications
        destination_ws.column_dimensions = dim_holder

    return


def resize_summary_columns(writer):
    """ Resize columns width in Summary sheet. """

    destination_ws = writer.sheets['summary']
    dim_holder = DimensionHolder(worksheet=destination_ws)

    # filename columns (A = 1)
    dim_holder[get_column_letter(1)] = ColumnDimension(destination_ws, min=1, max=1, width=75)

    # analysis summary columns
    last_col = destination_ws.max_column
    for col in range(2, last_col+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, bestFit=True)
    destination_ws.column_dimensions = dim_holder

    return


#######################
#        ROWS      #
#######################

def resize_rows(writer, sheets):
    """ Resize row height for each .xlsx sheet. """

    for sheet in sheets:
        destination_ws = writer.sheets[sheet]
        row_dim = RowDimension(destination_ws, index=1)
        row_dim.height = 20
    return


#######################
#        MAIN      #
#######################


def format_file(writer, filename, sheets, n_channels, progress_bar=None):
    """ Handles formatting of colums width and row height all at once """

    if progress_bar:
        progress_bar = st.progress(0, text=f"Formatting final table: [0/2]")

    workbook = writer.book

    # resize columns
    resize_summary_columns(writer)
    resize_analysis_columns(writer, sheets, n_channels)
    if progress_bar:
        progress_bar.progress(1/2, text=f"Formatting final table: [1/2]")

    # resize rows
    resize_rows(writer, sheets)
    if progress_bar:
        progress_bar.progress(2/2, text=f"Formatting final table: [2/2]")

    workbook.save(filename)
    return
