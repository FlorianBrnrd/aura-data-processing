import os
from copy import copy

import openpyxl
import streamlit as st
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


#######################
#   UTILS FUNCTIONS   #
#######################


def copy_cell_style(template_cell, destination_cell):
    destination_cell.font = copy(template_cell.font)
    destination_cell.border = copy(template_cell.border)
    destination_cell.fill = copy(template_cell.fill)
    destination_cell.number_format = copy(template_cell.number_format)
    destination_cell.protection = copy(template_cell.protection)
    destination_cell.alignment = copy(template_cell.alignment)
    return


def copy_cell_value(template_cell, destination_cell):
    destination_cell.value = template_cell.value
    return


def copy_cells(template_ws, destination_ws, template_col, destination_col, row, copy_value, copy_style):
    # reading cell value from template file
    template_cell = template_ws.cell(row=row, column=template_col)

    # destination cell
    destination_cell = destination_ws.cell(row=row, column=destination_col)

    # copy template cell value
    if copy_value:
        copy_cell_value(template_cell, destination_cell)

    # copy template cell style
    if copy_style and template_cell.has_style:
        copy_cell_style(template_cell, destination_cell)

    return


def copy_from_template(template_ws, destination_ws, start_row, end_row, template_start_col, template_end_col,
                       destination_start_col, destination_end_col, copy_value=True, copy_style=True):
    for row in range(start_row, end_row + 1):
        # copy to different columns
        if (template_start_col != destination_start_col) or (template_end_col != destination_end_col):
            for destination_col, template_col in zip(range(template_start_col, template_end_col + 1),
                                                     range(destination_start_col, destination_end_col + 1)):
                copy_cells(template_ws, destination_ws, template_col=template_col, destination_col=destination_col,
                           row=row, copy_value=copy_value, copy_style=copy_style)
        # copy to same columns
        else:
            for col in range(template_start_col, template_end_col + 1):
                copy_cells(template_ws, destination_ws, template_col=col, destination_col=col, row=row,
                           copy_value=copy_value, copy_style=copy_style)
    return


def add_separator(destination_ws, end_row, color='FF000000'):
    for col_range in range(1, 6):
        cell_title = destination_ws.cell(end_row + 1, col_range)
        cell_title.fill = PatternFill(fgColor=color, fill_type="solid")


def add_conditional_formatting(ws, col_start, col_end, row_start, row_end, end_color='000000'):
    columns_range = f'{col_start}{row_start}:{col_end}{row_end}'
    fill_color = PatternFill(fill_type='solid', fgColor='00' + end_color)
    format_rule = CellIsRule(operator='greaterThan', formula=['0'], fill=fill_color)
    ws.conditional_formatting.add(columns_range, format_rule)


def get_channel_color(index):
    channel_mappings = {
        1: 'e99bb5',
        2: 'f9dfe5',
        3: 'f2e4e9',
        4: 'fff2ef',
        5: 'ffefe7',
        6: 'e1eef1',
        7: 'B2DFDB',
        8: 'E8F6F3',
        9: 'ddf2f1',
        10: 'def4fb',
        11: 'c2ecff',
        12: 'd5dbee',
        13: 'e6e9f1',
        14: 'cfd0d3',
        15: 'eaebed',
    }

    return channel_mappings.get(index, None)


##############################
#   PARSE SUMMARY TEMPLATE   #
##############################


def parse_summary_template(writer, filename, file_channels, summary_template: str):
    workbook = writer.book
    destination_ws = writer.sheets['summary']
    max_rowdata = destination_ws.max_row

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), summary_template)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    # copy template analysis
    n_channels = max([len(channels) for channels in file_channels.values()])

    copy_from_template(template_ws, destination_ws, start_row=1, end_row=max_rowdata,
                       template_start_col=2, template_end_col=4 * n_channels,
                       destination_start_col=2, destination_end_col=4 * n_channels)

    # Add channels name to summary header
    chans = {}
    for sample, channels in file_channels.items():
        for channel, channel_name in channels.items():
            chans[channel] = channel_name

    for n, (channel, name) in enumerate(chans.items(), start=0):
        col = (4 * n) + 2
        cell = destination_ws.cell(1, col)
        cell.value = f'{channel}:'
        col = (4 * n) + 3
        cell = destination_ws.cell(1, col)
        cell.value = name

    # save file
    workbook.save(filename)
    return


##############################
#   PARSE ANALYSIS TEMPLATE  #
##############################


def copy_columns_style(writer, filename, sheets, sheet_template: str):
    """
    Copy style from template for data columns (merged from input files)
    """

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), sheet_template)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    workbook = writer.book

    for sheet in sheets:

        destination_ws = writer.sheets[sheet]
        max_rowdata = max(destination_ws.max_row, 50)
        max_coldata = destination_ws.max_column
        min_coldata = destination_ws.min_column

        # copy template RNAscope header
        copy_from_template(template_ws, destination_ws, start_row=1, end_row=2,
                           template_start_col=min_coldata, template_end_col=max_coldata,
                           destination_start_col=min_coldata, destination_end_col=max_coldata)

        # format RNAscope columns
        copy_from_template(template_ws, destination_ws, start_row=3, end_row=max_rowdata,
                           template_start_col=min_coldata, template_end_col=max_coldata,
                           destination_start_col=min_coldata, destination_end_col=max_coldata,
                           copy_value=False, copy_style=True)

        for i, col in enumerate(range(min_coldata + 1, max_coldata + 1), start=1):
            color = get_channel_color(i)
            add_conditional_formatting(ws=destination_ws, row_start=4, row_end=max_rowdata,
                                       col_start=get_column_letter(col),
                                       col_end=get_column_letter(col),
                                       end_color=color)

        # copy blank column
        copy_from_template(template_ws=template_ws, destination_ws=destination_ws, start_row=1, end_row=max_rowdata,
                           template_start_col=6, template_end_col=6,
                           destination_start_col=max_coldata + 1, destination_end_col=max_coldata + 1,
                           copy_value=False, copy_style=True)

    workbook.save(filename)
    return


def parse_analysis_template(writer, filename, sheets, file_channels, analysis_type, add_copositivity, sheet_template: str):
    """ Copy template to analyse image data - e.g. %Cell, H-score, etc """

    analysis_end = {}

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), sheet_template)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    workbook = writer.book

    for sheet in sheets:

        destination_ws = writer.sheets[sheet]

        # set where to stop parsing analysis row
        channels = file_channels[sheet]
        n_channels = len(channels)
        end_row = 2 + (8 * n_channels) if analysis_type == 'Count' else 2 + (5 * n_channels)

        # copy template analysis
        copy_from_template(template_ws, destination_ws, start_row=1, end_row=end_row,
                           template_start_col=1, template_end_col=6,
                           destination_start_col=1, destination_end_col=6)

        # copy image data channel header into analysis part
        for n in range(n_channels):
            if analysis_type == 'Count':
                destination_ws.cell(3 + n * 8, 1).value = destination_ws.cell(2, n + 8).value
                destination_ws.cell(3 + n * 8, 2).value = destination_ws.cell(3, n + 8).value
            elif analysis_type == 'Area':
                destination_ws.cell(3 + n * 5, 1).value = destination_ws.cell(2, n + 8).value
                destination_ws.cell(3 + n * 5, 2).value = destination_ws.cell(3, n + 8).value
            else:
                # do nothing
                continue

        analysis_end[sheet] = end_row

        if not add_copositivity:
            add_separator(destination_ws, end_row)

    workbook.save(filename)
    return analysis_end


def rename_channels_from_settings(writer, file_channels):
    """ Attribute correct channel number based on settings file """

    for file, channels in file_channels.items():
        destination_ws = writer.sheets[file]

        # remapping based on channel name
        for n, (channel_number, channel) in enumerate(channels.items(), start=8):
            if channel == destination_ws.cell(row=3, column=n).value:
                destination_ws.cell(row=2, column=n).value = channel_number


##############################
#       MAIN FUNCTION        #
##############################


def main_parsing(writer, filename, sheets, file_channels, add_copositivity, analysis_type, summary_template,
                 sheet_template, progress_bar=None):
    # Initialize progress bar
    steps = 3
    step = 0
    if progress_bar:
        progress_bar = st.progress(step, text=f"Parsing templates: [{step}/{steps}]")

    # AURA-macro data table
    copy_columns_style(writer=writer, filename=filename, sheets=sheets, sheet_template=sheet_template)
    rename_channels_from_settings(writer=writer, file_channels=file_channels)
    if progress_bar:
        step += 1
        progress_bar.progress(step / steps, text=f"Parsing templates: [{step}/{steps}]")

    # Analysis template
    analysis_end = parse_analysis_template(writer=writer, filename=filename, sheets=sheets,
                                           add_copositivity=add_copositivity, file_channels=file_channels,
                                           sheet_template=sheet_template, analysis_type=analysis_type)
    if progress_bar:
        step += 1
        progress_bar.progress(step / steps, text=f"Parsing templates: [{step}/{steps}]")

    # Summary template
    parse_summary_template(writer=writer, filename=filename, file_channels=file_channels,
                           summary_template=summary_template)

    if progress_bar:
        step += 1
        progress_bar.progress(step / steps, text=f"Parsing templates: [{step}/{steps}]")

    return analysis_end
