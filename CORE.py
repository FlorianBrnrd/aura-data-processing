# standard libraries
import os
import glob
import logging
import io
import re
from itertools import combinations
from copy import copy
from pathlib import Path

# third-party libraries
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension
from openpyxl.styles import PatternFill
from pydantic.v1.utils import deep_update



####################
#       utils      #
####################



def create_directory(directory_path):
    """ Creates a directory if it does not exist """
    out_path = Path(f'{directory_path}/')
    if not out_path.exists():
        os.makedirs(out_path, exist_ok=True)
    return


def create_xlsx_file(input_ID, output_folder='.'):
    """ Creates an .xlsx file that can be manipulated """
    file_name = os.path.join(output_folder, f"{input_ID}.xlsx")
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    return writer, file_name


def copy_cell_style(template_cell, destination_cell):
    destination_cell.font = copy(template_cell.font)
    destination_cell.border = copy(template_cell.border)
    destination_cell.fill = copy(template_cell.fill)
    destination_cell.number_format = copy(template_cell.number_format)
    destination_cell.protection = copy(template_cell.protection)
    destination_cell.alignment = copy(template_cell.alignment)
    return


def copy_cell_value(template_cell, destination_cell):
    destination_cell.value = template_cell.value
    return


def copy_cells(template_ws, destination_ws, template_col, destination_col, row, copy_value, copy_style):
    
    # reading cell value from template file
    template_cell = template_ws.cell(row=row, column=template_col)

    # destination cell
    destination_cell = destination_ws.cell(row=row, column=destination_col)
    
    # copy template cell value
    if copy_value:
        copy_cell_value(template_cell, destination_cell)
        
    # copy template cell style
    if copy_style and template_cell.has_style:
        copy_cell_style(template_cell, destination_cell)
    
    return


def copy_from_template(template_ws, destination_ws, start_row, end_row, template_start_col, template_end_col, destination_start_col, destination_end_col, copy_value=True, copy_style=True):

    for row in range(start_row, end_row+1):
        # copy to different columns
        if (template_start_col != destination_start_col) or (template_end_col != destination_end_col):
            for destination_col, template_col in zip(range(template_start_col, template_end_col+1), range(destination_start_col, destination_end_col+1)):
                copy_cells(template_ws, destination_ws, template_col=template_col, destination_col=destination_col, row=row, copy_value=copy_value, copy_style=copy_style)
        # copy to same columns
        else:
            for col in range(template_start_col, template_end_col+1):
                copy_cells(template_ws, destination_ws, template_col=col, destination_col=col, row=row, copy_value=copy_value, copy_style=copy_style)
    return



#####################
#  FILE PROCESSING  #
#####################


def get_channels_from_settings_file(uploaded_file):
    """ Parse channels present in settings file and returns it as a list """
    channels = []
    
    for line in uploaded_file:

        if type(line) is bytes:
            line = line.decode('utf-8', 'backslashreplace')
            
        if line.startswith('Channel'):
            match = re.match(r'(Channel [\d]: )(.+)$', line)
            channel = str(match.group(2)).strip()
            channels.append(channel)

    return channels


def build_files_attributes_dict(files_dict, channels_list, error_space):
    """ Input: a {filename : filedata} dictionnary"""
    
    # generate progress bar
    samples = len(files_dict)
    i = 0
    progress_bar = st.progress(0, text=f"Processing input files: [{i}/{samples}]")
    
    # processing files
    errors = []
    files_attributes = {}
    for filename, filedata in files_dict.items():
        
        # use regex to process file names
        matches = re.match(r'^(.+)_(.+).csv$', filename)
        sample = str(matches.group(1))
        channel = str(matches.group(2))
        
        if channel not in channels_list:
            errors.append((filename, channel))
             
        # make dictionnary from values
        file_dict = {sample: {channel: filedata}}

        # update main dictionnary
        files_attributes = deep_update(files_attributes, file_dict)
        
        # updating progress bar
        i += 1
        progress_bar.progress(i/samples, text=f'Processing input files: [{i}/{samples}]')
        
    
    if errors:
        string = [f'Found unknown channel [**{channel}**] in file: **{filename}**\n\n' for (filename, channel) in errors]
        string = ''.join(string)
        settings_chans = '| '.join(channels_list)
        settings_chans = f'File **Analysis_settings.txt** specify the following channels: {settings_chans}'
        string = string + settings_chans
        error_space.error(string)
        

    # sort dict
    files_attributes = {key: dict(sorted(files_attributes[key].items())) for key in sorted(files_attributes)}
        
    return files_attributes


def merge_image_channels(files_attributes, writer, file_name):
    """ For each image sample, merge corresponding channels together """

    # generate progress bar
    samples = len(files_attributes)
    i = 0
    progress_bar = st.progress(0, text=f"Merging image channels: [{i}/{samples}]")
    
    # initizalize variables
    counter = 3
    sheets = []
    data = {}
    file_channels = {}
    
    #create summary sheet in first position
    workbook = writer.book
    workbook.create_sheet('summary')
    summary_sheet = writer.sheets['summary']
    
    # Loop over samples
    for sample, channels_per_image in files_attributes.items():
        
        sheets.append(sample)
        
        # write filename to summary file and increment counter for next iteration
        summary_sheet[f"A{counter}"] = sample
        counter = counter + 1
        
        # store separate channel dataframes in list
        image_dfs = []
        for channel, filedata in channels_per_image.items():
            df_length = len(filedata)
            slices = [f'Slice_{i}' for i in range(1, df_length+1)]
            filedata['Slice'] = slices
            filedata = filedata.set_index('Slice')[['Count']]
            filedata.columns = [channel]
            image_dfs.append(filedata)
            
        # concat dataframes from same sample+channel, reorder columns based on names and store in dictionnary
        df = pd.concat(image_dfs, axis=1)
        df = df.reindex(sorted(df.columns), axis=1)
        data[sample] = df
        
        # build dictionnary from channel names
        file_channels[sample] = {f'Channel {i} (C{i})':col for i, col in enumerate(df.columns, start=1)}
        
        # write dataframe in a seperate sheet (one per channel)
        df.to_excel(writer, sheet_name=sample, startrow=2, startcol=6, index=True, header=True, na_rep='NaN')
        
        # updating progress bar
        i += 1
        progress_bar.progress(i/samples, text=f'Merging image channels: [{i}/{samples}]')
            
    # save file
    workbook.save(file_name)

    return sheets, data, file_channels


########################
#  RNASCOPE ANALYSIS   #
########################


def format_rnascope(writer, filename, sheets, template_file='template_sheet_final.xlsx'):
    """
    Copy style from template for data columns (merged from input files)
    """

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), template_file)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    workbook = writer.book

    for sheet in sheets:
        
        destination_ws = writer.sheets[sheet]
        max_rowdata = max(destination_ws.max_row, 50)
        max_coldata = destination_ws.max_column
        min_coldata = destination_ws.min_column
        
        # copy template RNAscope header       
        copy_from_template(template_ws, destination_ws, start_row=1, end_row=2,
                           template_start_col=min_coldata, template_end_col=max_coldata, 
                           destination_start_col=min_coldata, destination_end_col=max_coldata)

        
        # format RNAscope columns
        copy_from_template(template_ws, destination_ws, start_row=3, end_row=max_rowdata, 
                           template_start_col=min_coldata, template_end_col=max_coldata, 
                           destination_start_col=min_coldata, destination_end_col=max_coldata,
                           copy_value=False)

        # copy blank column
        copy_from_template(template_ws, destination_ws, start_row=1, end_row=max_rowdata, 
                           template_start_col=23, template_end_col=23, 
                           destination_start_col=max_coldata+1, destination_end_col=max_coldata+1, 
                           copy_value=False)
       
        
    workbook.save(filename)
    
    return


def resize_columns_rnascope(writer, filename, sheets):
    
    workbook = writer.book
    
    for sheet in sheets:
        
        destination_ws = writer.sheets[sheet]
        max_coldata = destination_ws.max_column
        min_coldata = destination_ws.min_column
        
        # Resize columns 
        dim_holder = DimensionHolder(worksheet=destination_ws)
        
        for col in range(min_coldata, max_coldata+1):
            dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, bestFit=True)
        
        destination_ws.column_dimensions = dim_holder
            
    # saving the destination excel file
    workbook.save(filename)
    
    return


########################
#   ANALYSIS TEMPLATE   #
########################


def parse_analysis(writer, filename, sheets, file_channels, template_file='template_sheet_final.xlsx'):
    """ Copy template to analyse image data - e.g. %Cell, H-score, etc """

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), template_file)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    workbook = writer.book
    
    for sheet in sheets:
        
        destination_ws = writer.sheets[sheet]
    
        # set where to stop parsing analysis row
        channels = file_channels[sheet]
        n_channels = len(channels)
        end_row = 2 + (8 * n_channels)
        
        # copy template analysis        
        copy_from_template(template_ws, destination_ws, start_row=1, end_row=end_row, 
                           template_start_col=1, template_end_col=6, 
                           destination_start_col=1, destination_end_col=6)
        
        # add a black row for separation
        for col_range in range(1, 6):
            cell_title = destination_ws.cell(end_row+1, col_range)
            cell_title.fill = PatternFill(fgColor='FF000000', fill_type="solid")

    workbook.save(filename)
    return 


def resize_columns_analysis(writer, filename, sheets):
    
    workbook = writer.book
    
    for sheet in sheets:
        
        destination_ws = writer.sheets[sheet]
        dim_holder = DimensionHolder(worksheet=destination_ws)
        
       # column A
        col = 1
        dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=40)
        
        # columns covering analysis (B to E)
        for col in range(2, 6):
            dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, bestFit=True)
        
        # column used as separator (F)
        col = 6
        dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, bestFit=True)
        
        destination_ws.column_dimensions = dim_holder

    workbook.save(filename)
    return
    

########################
#   SUMMARY TEMPLATE   #
########################


def parse_summary_template(writer, filename, sheets, file_channels=None, template_file='template_summary_final.xlsx'):
    
    workbook = writer.book
    destination_ws = writer.sheets['summary']
    max_rowdata = destination_ws.max_row
    
    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), template_file)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    # copy template analysis
    n_channels = max([len(channels) for channels in file_channels.values()])
    
    copy_from_template(template_ws, destination_ws, start_row=1, end_row=max_rowdata, 
                       template_start_col=2, template_end_col=4*n_channels, 
                       destination_start_col=2, destination_end_col=4*n_channels)

    # Add channels name to summary header
    chans = {}
    for sample, channels in file_channels.items():
        for channel, channel_name in channels.items():
            chans[channel] = channel_name
    
    for n, (channel, name) in enumerate(chans.items(), start=0):
        col = (4*n)+2
        cell = destination_ws.cell(1, col)
        cell.value = f'{channel}:'
        col = (4*n)+3
        cell = destination_ws.cell(1, col)
        cell.value = name
        
    # save file
    workbook.save(filename)
    return


def resize_columns_summary(writer, filename):
    
    workbook = writer.book
    destination_ws = writer.sheets['summary']
    
    dim_holder = DimensionHolder(worksheet=destination_ws)
    
    # filename columns
    col=1
    dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=50)

    # analysis summary columns
    last_col = destination_ws.max_column
    for col in range(2, last_col+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, bestFit=True)
    
    destination_ws.column_dimensions = dim_holder
    workbook.save(filename)
    return
    


#######################
#     FORMAT ROWS     #
#######################


def resize_row_height(writer, filename, sheets):

    workbook = writer.book
    
    for sheet in sheets:
        destination_ws = writer.sheets[sheet]
        row_dim = RowDimension(destination_ws, index=1) 
        row_dim.height = None 
    
    workbook.save(filename)
    return



####################
#   COPOSITIVITE   #
####################





####################
# MAIN PROCESSING #
####################


def process(input_ID, files_attributes, output_folder='.'):
    
    # create file
    writer, filename = create_xlsx_file(input_ID, output_folder='.')
    
### PROCESSING
    
    # 1) merge RNAscope tables
    sheets, data, file_channels = merge_image_channels(files_attributes, writer, filename)

    
### FORMATTING
    
    # generate progress bar
    progress_bar = st.progress(0, text=f"Formatting final table: [0/4]")
    
    # 2) Format RNAscope data table
    format_rnascope(writer=writer, filename=filename, sheets=sheets)
    progress_bar.progress(25, text=f"Formatting final table: [1/4]")
    
    # 3) Parse analysis
    parse_analysis(writer=writer, filename=filename, sheets=sheets, file_channels=file_channels)
    progress_bar.progress(50, text=f"Formatting final table: [2/4]")

    # 4) Parse and format summary sheet
    parse_summary_template(writer=writer, filename=filename, sheets=sheets, file_channels=file_channels)
    progress_bar.progress(75, text=f"Formatting final table: [3/4]")
    
    # 5) co-positivite
    #

    # 6) resize columns
    resize_columns_analysis(writer=writer, filename=filename, sheets=sheets)
    resize_columns_rnascope(writer=writer, filename=filename, sheets=sheets)
    #
    resize_columns_summary(writer=writer, filename=filename)
    
    # 7) resize rows
    resize_row_height(writer=writer, filename=filename, sheets=sheets)
    
    progress_bar.progress(100, text=f"Formatting final table: [4/4]")
    
    return
