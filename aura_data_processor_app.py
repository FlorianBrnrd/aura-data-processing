# standard libraries
import os
import io
import zipfile
from copy import copy
from pathlib import Path

# third party libraries
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from pydantic.v1.utils import deep_update



########################
#      USER INPUT      #
########################


def get_input_configuration():
    
    st.subheader('Input format', anchor=False)
    input_format = st.radio("input format", [".xls Files", ".zip Folder"], horizontal=True, label_visibility="hidden")

    if input_format == '.zip Folder':
        st.info('Input a single **.zip folder** containing the **.xls files** obtained by running the **AURA macro** on your images')

    elif input_format == '.xls Files':
        st.info('Input the **.xls files** obtained by running the **AURA macro** on your images')
    else:
        st.error('An error happened - please try again')
    
    st.write('###')
    return input_format


def get_uploaded_files(input_format):

    """ 
    Configure file_uploader based on user-defined input format
    """

    st.subheader('Input Files', anchor=False)

    if input_format == '.zip Folder':
        uploaded_files = st.file_uploader('Choose .zip folder to process', type='zip', accept_multiple_files=False)
    elif input_format == '.xls Files':
        uploaded_files = st.file_uploader('Choose .xls files to process', type='xls', accept_multiple_files=True)
    else:
        uploaded_files = None
    
    return uploaded_files


#########################
#   PROCESSING INPUT    #
#########################


@st.cache_data(show_spinner=False)
def unpack_zip_folder(input_zip):

    """ 
    Read .xls files contained into the input .zip folder
    """

    zipped_files = {}

    # Open zip file
    with zipfile.ZipFile(input_zip, "r") as zip_file:

        # list all files (path) in zip file
        for file_name in zip_file.namelist():

            # loop over files
            with zip_file.open(file_name, "r") as file_data:
                # get file name from path
                name = Path(file_name).name
                # check file is an .xls file and not something else
                if not (name.startswith('.') or name.startswith('_')) and (name.endswith('.xls')):
                    # get file content
                    data = io.BytesIO(file_data.read())
                    # create dict from file name and file content
                    zipped_files[name] = data.getvalue()

    return zipped_files

 
@st.cache_data(show_spinner=False)
def preprocess_zipped_sample_files(zipped_list):

    """ 
    Extract "sample", "channel" and "image number" from .xls files inside the .zip folder
    """

    # generate progress bar
    total_files = len(zipped_list)
    progress_bar = st.progress(0, text=f"Parsing files names: [0/{total_files}]")

    # Create dict to store files attributes
    files_dict = {}    

    # loop over files
    for i, (fname, fdata)  in enumerate(zipped_list.items(), start=1):

        # update progress bar
        progress_bar.progress(i/total_files, text=f'Parsing files names: [{i}/{total_files}]')

        # split file name
        f = fname.split('_')

        # Extract sample, channel and image from file name
        sample = '_'.join(f[:-2])
        channel = f[-1].replace('.xls','')
        image = int(f[-2])

        # make dictionnary from files attributes
        file_dict = {sample: { channel: {image: fdata}}}

        # update main dictionnary
        files_dict = deep_update(files_dict, file_dict)

    return files_dict


@st.cache_data(show_spinner=False)
def preprocess_xls_sample_files(files_list):
    
    """ 
    Extract "sample", "channel" and "image number" from .xls input files names
    """

    # generate progress bar
    total_files = len(files_list)
    progress_bar = st.progress(0, text=f"Parsing files names: [0/{total_files}]")

    # Create dict to store files attributes
    files_dict = {}    

    # loop over files
    for i, file in enumerate(files_list, start=1):

        # update progress bar
        progress_bar.progress(i/total_files, text=f'Parsing files names: [{i}/{total_files}]')

        # file data
        bytes_data = file.getvalue()
        
        # split file name
        fname = file.name
        f = fname.split('_')

        # Extract sample, channel and image from file name
        sample = '_'.join(f[:-2])
        channel = f[-1].replace('.xls','')
        image = int(f[-2])

        # make dictionnary from files attributes
        file_dict = {sample: { channel: {image: bytes_data}}}

        # update main dictionnary
        files_dict = deep_update(files_dict, file_dict)

    return files_dict


########################
#    MERGING TABLES    #
########################


@st.cache_data(show_spinner=False)
def merge_files(files_dict, output_folder='.'):
  
    """ 
    For each sample, we create one sheet per channel.
    For each channel, images tables gets merged together.
    """

    # Generate progress bar
    total_sample = len(files_dict)
    progress_bar = st.progress(0, text=f"Processing and merging files: [0/{total_sample}]")

    # Loop over samples
    for i, (sample, channels) in enumerate(files_dict.items(), start=1):

        # update progress bar
        progress_bar.progress(i/total_sample, text=f'Processing and merging files: [{i}/{total_sample}]')

        # For each sample, create a file
        file_name = os.path.join(output_folder, f"{sample}.xlsx")
        writer = pd.ExcelWriter(file_name, engine='openpyxl')

        # loop over channels
        for channel, images_per_channel in channels.items():

            # create dataframe
            channel_df = pd.DataFrame()

            for image, filepath in images_per_channel.items():

                # load data as dataframe
                filepath = io.StringIO(filepath.decode("utf-8"))
                
                df = pd.read_csv(filepath, sep='\t', index_col='Slice')
                df = pd.concat([df], keys=[f'img{image}'], names=[''], axis=1)

                # concat dataframes from same sample and channel together
                channel_df = pd.concat([channel_df, df], axis=1)

            # reorder top-level columns based on image number
            channel_df = channel_df.sort_index(axis=1, level=[0], ascending=[True], inplace=False)
            
            # write dataframe in a seperate sheet (one per channel)
            channel_df.to_excel(writer, sheet_name=channel, startrow=0 , startcol=5, index=True, header=True, na_rep='NaN')

            ### Formatting file

            # Get the xlsxwriter workbook and worksheet objects
            workbook  = writer.book
            worksheet = writer.sheets[channel]

            # Rename F2 cell 
            worksheet["F2"] = "Slice"

            # Remove empty row created by pandas multiindex
            worksheet.delete_rows(3)  

        ### Saving file
        workbook.save(file_name)

    return


########################
#   PARSING ANALYSIS   #
########################


@st.cache_data(show_spinner=False)
def add_analyse(files_dict, output_folder='.', template_file='analysis_template.xlsx'):
  
    """
    Open sample files generated by merge_files() 
    In each sheet, parse the analysis template allowing to measure "%cells" and "H-Score"
    based on image's count data
    """

    # generate progress bar
    total_sample = len(files_dict)
    progress_bar = st.progress(0, text=f"Appending results and saving files: [0/{total_sample}]")

    # Open the template file
    template_rel_path = os.path.join(os.path.dirname(__file__), template_file)
    template_wb = openpyxl.load_workbook(template_rel_path)
    template_ws = template_wb.worksheets[0]

    # calculate total number of rows and columns in template file
    max_row = template_ws.max_row
    max_col = template_ws.max_column

    # Loop over samples
    for i, (sample, channels) in enumerate(files_dict.items(), start=1):

        # update progress bar
        progress_bar.progress(i/total_sample, text=f'Appending results and saving files: [{i}/{total_sample}]')
 
        # loop over channels
        for channel in channels:

            # Open the destination excel file
            sample_file = os.path.join(output_folder, f"{sample}.xlsx")

            destination_wb = openpyxl.load_workbook(sample_file)
            destination_ws = destination_wb[channel]

            # copying the cell values from source excel file to destination excel file
            for i in range (1, max_row + 1):
                for j in range (1, max_col + 1):

                    # reading cell value from template file
                    template_cell = template_ws.cell(row = i, column = j)

                    # destination cell
                    destination_cell = destination_ws.cell(row = i, column = j)

                    # copy cell value
                    destination_cell.value = template_cell.value

                    # copy cell style
                    if template_cell.has_style:
                        destination_cell.font = copy(template_cell.font)
                        destination_cell.border = copy(template_cell.border)
                        destination_cell.fill = copy(template_cell.fill)
                        destination_cell.number_format = copy(template_cell.number_format)
                        destination_cell.protection = copy(template_cell.protection)
                        destination_cell.alignment = copy(template_cell.alignment)

                # Resize columns 
                dim_holder = DimensionHolder(worksheet=destination_ws)
                
                # columns covering analysis (A to D)
                for col in range(1, 5):
                    dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=15)
                
                # column used as separator (E)
                for col in range(5, 6):
                    dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=3)
                destination_ws.column_dimensions = dim_holder

                # columns covering images data (F to last one)
                for col in range(6, destination_ws.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(destination_ws, min=col, max=col, width=12)
               
            # saving the destination excel file
            destination_wb.save(str(sample_file))

    return


########################
#       WEB APP        #
########################

def app_settings():

    st.set_page_config(layout="wide", page_title='AURA Data Processor', page_icon='🔬')

    hide_streamlit_style = """
                <style>
                #MainMenu {visibility: hidden;}
                footer {visibility: hidden;}
                .css-1d391kg {padding-top: 3rem; padding-bottom: 1rem}
                .css-18e3th9 {padding-top: 3rem; padding-bottom: 1rem}
                header[data-testid="stHeader"] {background: none}
                </style>
                """

    st.markdown(hide_streamlit_style, unsafe_allow_html=True)


def sidebar_howto():

    with st.sidebar:
        st.write('# **How to :blue[use] ?**')
        with st.expander('See explanations', expanded=False):
            st.write('**1. Choose desired input format**')
            st.write('**2. Upload your files accordingly**')
            st.write('**3. Click on "Process files" to run the script**')
            st.write('**4. Follow script progression**')
            st.write('**5. Click on "Download" to get your results**')
        st.divider()
    
    return


def rnascope():

    uploaded_files = None
    
    # set app settings
    app_settings()

    # Show infos in sidebar
    sidebar_howto()

    # Main header
    st.header(':blue[AURA Data Processor]', anchor=False, divider='grey')

    # choose and filter file type to be processed
    input_format = get_input_configuration()

    # file uploader
    uploaded_files = get_uploaded_files(input_format)

    placeholder = st.empty()
    placeholder.button('Process files', disabled=True, key=12)

    if uploaded_files:
        
        process = placeholder.button('Process files', disabled=False, key=21, type="primary")

        if process:
            
            st.subheader('Progress', anchor=False)

            # create source dir to store files being processed
            out_path = Path('source_dir/')
            if not out_path.exists():
                os.makedirs(out_path, exist_ok=True)

            ## PROCESS INPUT ##
            if input_format == '.zip Folder':
                zipped_list = unpack_zip_folder(uploaded_files)
                processed_files = preprocess_zipped_sample_files(zipped_list)

            elif input_format == '.xls Files':
                processed_files = preprocess_xls_sample_files(uploaded_files)

            else:
                st.error('Files could not be processed - check input and retry')
                st.stop()

            if not processed_files or len(processed_files)<1:
                st.error('Files could not be processed - check input and retry')
                st.stop()
            
            ## MERGE TABLES ##
            merge_files(processed_files, output_folder=out_path)

            ## PARSING ANALYSIS TEMPLATE ##
            add_analyse(processed_files, output_folder=out_path)

            ## DOWNLOAD FILES ##
            st.subheader('Results', anchor=False)
            
            # Add output files into a zip folder
            with zipfile.ZipFile("directory.zip", mode="w") as archive:
                for file_path in out_path.iterdir():
                    # add files to zip folder
                    archive.write(file_path, arcname=file_path.name)
                    # clean files once added to zip folder
                    os.remove(file_path)

            # remove directory to free space 
            os.removedirs(out_path)

            # download zipped folder containing output files
            with open("directory.zip", "rb") as fp:
                download = st.download_button(label='Download', data=fp, file_name='results.zip', mime="application/zip", type="primary")

            # remove zip folder to free space
            os.remove("directory.zip")

    return

        
if __name__ == '__main__':

    rnascope()
